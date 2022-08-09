import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
import pyinputplus as pyip


while(True):
    n = pyip.inputNum('Enter an integer: ')

    if(isinstance(n, int)):
        break
    else:
        print(str(n) + " is not an integer.")

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = 'Table'
boldFont = Font(bold=True)

for i in range(1, n + 1):
    sheet['A' + str(i + 1)] = i
    sheet['A' + str(i + 1)].font = boldFont

for i in range(1, n + 1):
    sheet[str(get_column_letter(i + 1)) + '1'] = i
    sheet[str(get_column_letter(i + 1)) + '1'].font = boldFont

    for j in range(2, sheet.max_row + 1):
        sheet[str(get_column_letter(i + 1)) + str(j)] = sheet[str(get_column_letter(i + 1) + '1')].value * sheet['A' + str(j)].value

wb.save('multiplication-chart.xlsx')