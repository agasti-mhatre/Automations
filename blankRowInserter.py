import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pyinputplus as pyip

while(True):
    n = pyip.inputNum("Enter the starting row: ")

    if(isinstance(n, int)):
        break
    else:
        print(str(n) + " is not an integer.")

while(True):
    m = pyip.inputNum("Enter how many rows of space you would like: ")

    if(isinstance(m, int)):
        break
    else:
        print(str(m) + " is not an integer.")

n = int(n)
m = int(m)

wb = openpyxl.load_workbook('myProduce.xlsx')
sheet = wb.active

#get the values of the cells of the rows that will be needed after the space
rows_after_space = []
for i in range(n, sheet.max_row + 1):
    rows_after_space.append(list())
    for j in range(1, sheet.max_column + 1):
        rows_after_space[i - n].append(sheet[get_column_letter(j) + str(i)].value)

#create the spaces
for row in range(n, n + sheet.max_row):
    for column in range(1, sheet.max_column + 1):
        sheet[str(get_column_letter(column)) + str(row)] = ''

#insert the rows after the spaces
list_row = 0
for row in range(n + m, n + m + len(rows_after_space)):
    element = 0
    for column in range(1, sheet.max_column + 1):
        sheet[str(get_column_letter(column)) + str(row)] = rows_after_space[list_row][element]
        element += 1
    list_row += 1


wb.save('myProduce_new.xlsx')