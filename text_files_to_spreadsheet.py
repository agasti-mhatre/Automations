import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


#start a new workbook and work in sheet
excel_file = openpyxl.Workbook()
sheet = excel_file.active
sheet.title = 'text'


#find current directoy
p = Path.cwd()


column = 1
#iterate through each text file in the directory and retrieve contents
for file in list(p.glob('*.txt')):
    one_file = open(file)
    content = one_file.read()
    row = 1
    content_string = ''
    for letter in content:
        content_string += letter
        if letter == '\n':
            sheet[str(get_column_letter(column)) + str(row)] = content_string
            content_string = ''
            row += 1
    
    if content_string != '':
        sheet[str(get_column_letter(column)) + str(row)] = content_string
        content_string = ''
        
    column += 1
    
        


excel_file.save('text_to_excel.xlsx')

