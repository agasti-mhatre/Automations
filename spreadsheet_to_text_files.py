import os
from pathlib import Path
from types import NoneType
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


excel_file = openpyxl.load_workbook('excel_to_text.xlsx')
sheet = excel_file.active

p = Path.cwd()

for file in list(p.glob('*.txt')):
    os.remove(file)

for column in range(1, sheet.max_column + 1):
    new_text_file = open('column ' + str(column) + '.txt', 'a')
    row = 1
    while(not(isinstance(sheet[str(get_column_letter(column)) + str(row)].value, NoneType))):
        new_text_file.write(sheet[str(get_column_letter(column)) + str(row)].value)
        row += 1

    new_text_file.close()