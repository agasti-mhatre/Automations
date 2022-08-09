import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook('items_sold.xlsx')
sheet = wb.active

#get all of the columns and row data
value_of_cell = list()
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1): 
        value_of_cell.append({sheet[str(get_column_letter(column)) + str(row)].value: {str(get_column_letter(column)): str(row)}})


#invert the excel sheet
new_wb = openpyxl.Workbook()
sheet_new = new_wb.active
sheet_new.title = 'items'

for dict in value_of_cell:
    cell_value = str(list(dict.keys())[0])
    old_column = list((list(dict.values())[0]).keys())[0]
    old_row = list((list(dict.values())[0]).values())[0]

    sheet_new[str(get_column_letter(int(old_row))) + str(column_index_from_string(old_column))] = str(cell_value)     



new_wb.save('items_sold_inverted.xlsx')